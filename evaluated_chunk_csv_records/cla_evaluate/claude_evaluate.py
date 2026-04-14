"""
RAG Chunking Strategy Benchmark
================================
Evaluates 7 chunking strategies × 4 retrieval techniques across all IR metrics.
Outputs:
  - raw_results.csv         : per-question scores
  - chunk_stats.csv         : chunk quality stats per method
  - summary.csv             : aggregated metrics per method+technique
  - benchmark_report.xlsx   : full formatted Excel workbook with all sheets + verdict
"""

import re
import warnings
import numpy as np
import pandas as pd
from collections import defaultdict
from scipy import stats
from tqdm import tqdm

from sentence_transformers import SentenceTransformer, CrossEncoder
from qdrant_client import QdrantClient
from sklearn.metrics.pairwise import cosine_similarity
from rank_bm25 import BM25Okapi
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

warnings.filterwarnings("ignore")

# ============================================================
# CONFIG
# ============================================================
CSV_PATH       = "/home/root473/Documents/POC/ChunkArena/Golden_dataset/Banking_system.csv"
FINAL_K        = 5
RETRIEVAL_K    = 50        # candidates before reranking
HYBRID_CAND_K  = 100       # candidates for BM25 + dense in hybrid
SOFT_THRESHOLD = 0.85      # semantic similarity threshold for relevance
CHUNK_METHODS  = ["fixed_size", "overlapping", "sentence", "paragraph",
                  "recursive", "header", "semantic"]
TECHNIQUES     = ["dense", "hybrid", "dense_rerank", "hybrid_rerank"]

# ============================================================
# LOAD GOLD DATA
# ============================================================
print("Loading gold dataset...")
gold_df = pd.read_csv(CSV_PATH)
questions_data = []
for _, row in gold_df.iterrows():
    questions_data.append({
        "id": len(questions_data),
        "question": row["Question"],
        "gold_spans": [s.strip() for s in str(row["Gold_spans"]).split(";")]
    })
print(f"  Loaded {len(questions_data)} questions")

# ============================================================
# INIT MODELS
# ============================================================
print("Loading models...")
device   = "cpu"
embedder = SentenceTransformer('BAAI/bge-m3', device=device)
cross_encoder = CrossEncoder('cross-encoder/ms-marco-MiniLM-L-6-v2', device=device)
client   = QdrantClient(host="localhost", port=6333)
print("  Models ready")

# ============================================================
# CACHES
# ============================================================
embedding_cache  = {}
retrieval_cache  = {}
rerank_cache     = {}
relevance_cache  = {}

def get_embedding(text):
    if text not in embedding_cache:
        embedding_cache[text] = embedder.encode([text], convert_to_numpy=True)[0]
    return embedding_cache[text]

# ============================================================
# LOAD CHUNKS + BM25
# ============================================================
word_re = re.compile(r'\w+')

def get_all_chunks(collection_name):
    all_points = []
    offset = None
    while True:
        batch, next_offset = client.scroll(
            collection_name=collection_name,
            limit=1000,
            offset=offset,
            with_payload=True
        )
        all_points.extend(batch)
        if next_offset is None:
            break
        offset = next_offset
    ids   = [p.id for p in all_points]
    texts = [p.payload["text"] for p in all_points]
    return ids, texts

all_chunks_cache    = {}
all_chunks_text_dict = {}
bm25_models         = {}

print("Loading collections and building BM25 indexes...")
for method in CHUNK_METHODS:
    ids, texts = get_all_chunks(method)
    all_chunks_cache[method]     = (ids, texts)
    all_chunks_text_dict[method] = dict(zip(ids, texts))
    tokenized = [word_re.findall(t.lower()) for t in texts]
    bm25_models[method] = BM25Okapi(tokenized)
    print(f"  {method}: {len(texts)} chunks")

# ============================================================
# CHUNK QUALITY STATS (independent of retrieval)
# ============================================================
def chunk_stats(method):
    _, texts = all_chunks_cache[method]
    lengths  = [len(t.split()) for t in texts]
    boundary_count = sum(
        1 for t in texts if re.search(r'[.!?]["\']?\s*$', t.strip())
    )
    return {
        "method"            : method,
        "num_chunks"        : len(texts),
        "avg_words"         : round(np.mean(lengths), 2),
        "std_words"         : round(np.std(lengths), 2),
        "min_words"         : int(np.min(lengths)),
        "max_words"         : int(np.max(lengths)),
        "median_words"      : round(float(np.median(lengths)), 2),
        "boundary_ratio"    : round(boundary_count / len(texts), 4),
        "collection_redundancy": round(collection_redundancy(method), 4)
    }

def collection_redundancy(method, sample_size=300):
    _, texts = all_chunks_cache[method]
    sample   = texts[:sample_size]
    embs     = np.array([get_embedding(t) for t in sample])
    sim      = cosine_similarity(embs)
    n        = len(sim)
    vals     = [sim[i][j] for i in range(n) for j in range(i+1, n)]
    return float(np.mean(vals)) if vals else 0.0

# ============================================================
# RETRIEVAL
# ============================================================
def dense_search(query_text, query_emb, method, top_k):
    key = ("dense", method, query_text, top_k)
    if key in retrieval_cache:
        return retrieval_cache[key]
    res   = client.query_points(collection_name=method,
                                query=query_emb.tolist(), limit=top_k)
    ids   = [r.id for r in res.points]
    texts = [all_chunks_text_dict[method][i] for i in ids]
    retrieval_cache[key] = (ids, texts)
    return ids, texts

def hybrid_search(query, query_emb, method, top_k):
    key = ("hybrid", method, query, top_k)
    if key in retrieval_cache:
        return retrieval_cache[key]
    dense_ids, _ = dense_search(query, query_emb, method, HYBRID_CAND_K)
    bm25    = bm25_models[method]
    tokens  = word_re.findall(query.lower())
    scores  = bm25.get_scores(tokens)
    all_ids, _ = all_chunks_cache[method]
    bm25_ranked = sorted(zip(all_ids, scores), key=lambda x: x[1], reverse=True)[:HYBRID_CAND_K]
    fused   = defaultdict(float)
    for i, cid in enumerate(dense_ids):
        fused[cid] += 1 / (60 + i)
    for i, (cid, _) in enumerate(bm25_ranked):
        fused[cid] += 1 / (60 + i)
    final_ids = sorted(fused.keys(), key=lambda x: fused[x], reverse=True)[:top_k]
    texts     = [all_chunks_text_dict[method][i] for i in final_ids]
    retrieval_cache[key] = (final_ids, texts)
    return final_ids, texts

def rerank(query, ids, texts, top_k):
    if not texts:
        return [], []
    key = (query, tuple(ids))
    if key in rerank_cache:
        s_ids, s_texts = rerank_cache[key]
        return s_ids[:top_k], s_texts[:top_k]
    pairs   = [[query, t] for t in texts]
    scores  = cross_encoder.predict(pairs)
    idx     = np.argsort(scores)[::-1]
    s_ids   = [ids[i] for i in idx]
    s_texts = [texts[i] for i in idx]
    rerank_cache[key] = (s_ids, s_texts)
    return s_ids[:top_k], s_texts[:top_k]

# ============================================================
# RELEVANCE — hybrid exact + semantic fallback
# ============================================================
def is_relevant(chunk, gold_spans):
    key = (chunk, tuple(gold_spans))
    if key in relevance_cache:
        return relevance_cache[key]
    chunk_lower = chunk.lower()
    for span in gold_spans:
        if span.lower() in chunk_lower:
            relevance_cache[key] = True
            return True
    chunk_emb = get_embedding(chunk)
    for span in gold_spans:
        score = cosine_similarity([chunk_emb], [get_embedding(span)])[0][0]
        if score >= SOFT_THRESHOLD:
            relevance_cache[key] = True
            return True
    relevance_cache[key] = False
    return False

# ============================================================
# METRICS
# ============================================================
def hit_at_k(chunks, gold_spans, k):
    return int(any(is_relevant(c, gold_spans) for c in chunks[:k]))

def mrr_score(chunks, gold_spans):
    for i, c in enumerate(chunks, 1):
        if is_relevant(c, gold_spans):
            return 1.0 / i
    return 0.0

def precision_at_k(chunks, gold_spans, k):
    return sum(1 for c in chunks[:k] if is_relevant(c, gold_spans)) / k

def ndcg_at_k(chunks, gold_spans, k):
    dcg  = sum(1 / np.log2(i + 2) for i, c in enumerate(chunks[:k])
               if is_relevant(c, gold_spans))
    idcg = sum(1 / np.log2(i + 2) for i in range(min(len(gold_spans), k)))
    return dcg / idcg if idcg > 0 else 0.0

def recall_at_k(chunks, gold_spans, k):
    if not gold_spans:
        return 0.0
    found = sum(1 for span in gold_spans
                if any(is_relevant(c, [span]) for c in chunks[:k]))
    return found / len(gold_spans)

def avg_rank_score(chunks, gold_spans):
    for i, c in enumerate(chunks, 1):
        if is_relevant(c, gold_spans):
            return i
    return np.nan

def redundancy_score(chunks):
    if len(chunks) < 2:
        return 0.0
    embs = np.array([get_embedding(c) for c in chunks])
    sim  = cosine_similarity(embs)
    n    = len(sim)
    vals = [sim[i][j] for i in range(n) for j in range(i+1, n)]
    return float(np.mean(vals)) if vals else 0.0

def boundary_score(chunks):
    if not chunks:
        return 0.0
    return sum(1 for c in chunks if re.search(r'[.!?]["\']?\s*$', c.strip())) / len(chunks)

# ============================================================
# EVALUATION LOOP
# ============================================================
print("\nComputing chunk quality stats...")
stats_rows = []
for method in CHUNK_METHODS:
    print(f"  {method}...")
    stats_rows.append(chunk_stats(method))
stats_df = pd.DataFrame(stats_rows)

print("\nRunning evaluation...")
raw_results = []

for method in CHUNK_METHODS:
    print(f"\n  Method: {method}")
    for q in tqdm(questions_data, desc=f"    {method}"):
        query      = q["question"]
        gold_spans = q["gold_spans"]
        q_emb      = get_embedding(query)

        dense_ids, dense_texts = dense_search(query, q_emb, method, RETRIEVAL_K)
        hybrid_ids, hybrid_texts = hybrid_search(query, q_emb, method, RETRIEVAL_K)

        techniques = {
            "dense"         : (dense_ids[:FINAL_K],   dense_texts[:FINAL_K]),
            "hybrid"        : (hybrid_ids[:FINAL_K],  hybrid_texts[:FINAL_K]),
            "dense_rerank"  : rerank(query, dense_ids,  dense_texts,  FINAL_K),
            "hybrid_rerank" : rerank(query, hybrid_ids, hybrid_texts, FINAL_K),
        }

        for tech, (ids, texts) in techniques.items():
            h  = hit_at_k(texts, gold_spans, FINAL_K)
            m  = mrr_score(texts, gold_spans)
            p  = precision_at_k(texts, gold_spans, FINAL_K)
            n  = ndcg_at_k(texts, gold_spans, FINAL_K)
            rc = recall_at_k(texts, gold_spans, FINAL_K)
            ar = avg_rank_score(texts, gold_spans)
            rd = redundancy_score(texts)
            b  = boundary_score(texts)

            raw_results.append({
                "method"      : method,
                "technique"   : tech,
                "question_id" : q["id"],
                "question"    : query,
                "hit@k"       : h,
                "mrr"         : round(m,  4),
                "precision@k" : round(p,  4),
                "ndcg@k"      : round(n,  4),
                "recall@k"    : round(rc, 4),
                "avg_rank"    : ar,
                "redundancy"  : round(rd, 4),
                "diversity"   : round(1 - rd, 4),
                "boundary"    : round(b,  4),
            })

raw_df = pd.DataFrame(raw_results)
raw_df.to_csv("raw_results.csv", index=False)
stats_df.to_csv("chunk_stats.csv", index=False)
print("\nRaw results saved")

# ============================================================
# SUMMARY + SIGNIFICANCE
# ============================================================
summary_df = raw_df.groupby(["method", "technique"]).agg(
    hit_at_k       = ("hit@k",       "mean"),
    mrr            = ("mrr",         "mean"),
    precision_at_k = ("precision@k", "mean"),
    ndcg_at_k      = ("ndcg@k",      "mean"),
    recall_at_k    = ("recall@k",    "mean"),
    avg_rank       = ("avg_rank",    "mean"),
    redundancy     = ("redundancy",  "mean"),
    diversity      = ("diversity",   "mean"),
    boundary       = ("boundary",    "mean"),
    n_questions    = ("question_id", "count"),
).round(4).reset_index()

# composite score: weighted average of key metrics
W = {"ndcg_at_k": 0.30, "mrr": 0.25, "hit_at_k": 0.20,
     "recall_at_k": 0.15, "precision_at_k": 0.10}
summary_df["composite_score"] = sum(
    summary_df[m] * w for m, w in W.items()
).round(4)

summary_df = summary_df.sort_values("composite_score", ascending=False).reset_index(drop=True)
summary_df["rank"] = summary_df.index + 1
summary_df.to_csv("summary.csv", index=False)

# pairwise significance per technique
sig_rows = []
for tech in TECHNIQUES:
    sub = raw_df[raw_df["technique"] == tech]
    methods = CHUNK_METHODS
    for i, m1 in enumerate(methods):
        for m2 in methods[i+1:]:
            s1 = sub[sub["method"] == m1]["ndcg@k"].dropna()
            s2 = sub[sub["method"] == m2]["ndcg@k"].dropna()
            if len(s1) > 1 and len(s2) > 1:
                t_stat, p_val = stats.ttest_rel(s1, s2)
                sig_rows.append({
                    "technique"   : tech,
                    "method_a"    : m1,
                    "method_b"    : m2,
                    "mean_a"      : round(s1.mean(), 4),
                    "mean_b"      : round(s2.mean(), 4),
                    "t_stat"      : round(t_stat, 4),
                    "p_value"     : round(p_val,  4),
                    "significant" : "Yes" if p_val < 0.05 else "No",
                    "better"      : m1 if s1.mean() > s2.mean() else m2,
                })
sig_df = pd.DataFrame(sig_rows)

# ============================================================
# VERDICT: best method per technique + overall best
# ============================================================
verdict_rows = []
for tech in TECHNIQUES:
    sub  = summary_df[summary_df["technique"] == tech].copy()
    best = sub.sort_values("composite_score", ascending=False).iloc[0]
    verdict_rows.append({
        "technique"       : tech,
        "best_method"     : best["method"],
        "composite_score" : best["composite_score"],
        "hit@k"           : best["hit_at_k"],
        "mrr"             : best["mrr"],
        "ndcg@k"          : best["ndcg_at_k"],
        "recall@k"        : best["recall_at_k"],
        "precision@k"     : best["precision_at_k"],
    })
verdict_df = pd.DataFrame(verdict_rows)

overall_best = summary_df.iloc[0]
print(f"\n{'='*60}")
print(f"OVERALL BEST: {overall_best['method']} + {overall_best['technique']}")
print(f"  Composite score : {overall_best['composite_score']}")
print(f"  nDCG@{FINAL_K}         : {overall_best['ndcg_at_k']}")
print(f"  MRR             : {overall_best['mrr']}")
print(f"  Hit@{FINAL_K}          : {overall_best['hit_at_k']}")
print(f"{'='*60}")

# ============================================================
# EXCEL WORKBOOK
# ============================================================
print("\nBuilding Excel report...")

DARK_BLUE  = "1F3864"
MED_BLUE   = "2F5496"
LIGHT_BLUE = "BDD7EE"
HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
ALT_ROW    = "EBF3FB"
GREEN_GOOD = "C6EFCE"
RED_BAD    = "FFC7CE"
AMBER_MID  = "FFEB9C"
GOLD       = "FFD700"
WHITE      = "FFFFFF"

def header_font(size=11, bold=True, color=HEADER_FG):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def header_fill(color=HEADER_BG):
    return PatternFill("solid", fgColor=color)

def cell_fill(color):
    return PatternFill("solid", fgColor=color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, row_num, cols):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font      = header_font()
        c.fill      = header_fill()
        c.alignment = center()
        c.border    = thin_border()

def apply_data_rows(ws, dataframe, start_row=2, alt=True):
    for r_idx, row in enumerate(dataframe.itertuples(index=False), start=start_row):
        fill = cell_fill(ALT_ROW) if (alt and r_idx % 2 == 0) else None
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font      = body_font()
            c.alignment = center()
            c.border    = thin_border()
            if fill:
                c.fill = fill

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

wb = Workbook()
wb.remove(wb.active)

# ----------------------------------------------------------
# SHEET 1 — COVER / VERDICT
# ----------------------------------------------------------
ws_cover = wb.create_sheet("📊 Verdict")

ws_cover.column_dimensions["A"].width = 3
ws_cover.column_dimensions["B"].width = 28
for col in ["C","D","E","F","G","H"]:
    ws_cover.column_dimensions[col].width = 18

ws_cover.merge_cells("B2:H2")
c = ws_cover["B2"]
c.value     = "RAG Chunking Strategy Benchmark — Final Verdict"
c.font      = Font(name="Arial", size=18, bold=True, color=DARK_BLUE)
c.alignment = center()

ws_cover.merge_cells("B3:H3")
c = ws_cover["B3"]
c.value     = f"Evaluated {len(CHUNK_METHODS)} chunking methods × {len(TECHNIQUES)} retrieval techniques | {len(questions_data)} questions | K={FINAL_K}"
c.font      = Font(name="Arial", size=11, color="595959")
c.alignment = center()

ws_cover.row_dimensions[2].height = 36
ws_cover.row_dimensions[3].height = 20
ws_cover.row_dimensions[4].height = 14

# overall winner box
ws_cover.merge_cells("B5:H7")
c = ws_cover["B5"]
c.value = (f"🏆  OVERALL WINNER:  {overall_best['method'].upper()}  +  "
           f"{overall_best['technique'].upper()}   "
           f"(Composite Score: {overall_best['composite_score']:.4f})")
c.font      = Font(name="Arial", size=14, bold=True, color=WHITE)
c.fill      = header_fill(MED_BLUE)
c.alignment = center()
ws_cover.row_dimensions[5].height = 40
ws_cover.row_dimensions[6].height = 40
ws_cover.row_dimensions[7].height = 40

# verdict table header
hdrs = ["Technique", "Best Method", "Composite", "Hit@K", "MRR", "nDCG@K", "Recall@K", "Precision@K"]
apply_header_row(ws_cover, 9, hdrs)
ws_cover.row_dimensions[9].height = 22

medal = ["🥇","🥈","🥉","  "]
for r, row in verdict_df.iterrows():
    rn  = 10 + r
    vals = [
        row["technique"], row["best_method"],
        row["composite_score"], row["hit@k"],
        row["mrr"], row["ndcg@k"],
        row["recall@k"], row["precision@k"],
    ]
    fill_col = [GREEN_GOOD, GREEN_GOOD] + [None]*6
    for ci, val in enumerate(vals, 2):
        c = ws_cover.cell(row=rn, column=ci, value=val)
        c.font      = body_font(bold=(ci <= 3))
        c.alignment = center()
        c.border    = thin_border()
        if fill_col[ci-2]:
            c.fill = cell_fill(fill_col[ci-2])
    ws_cover.row_dimensions[rn].height = 20

# weight note
note_row = 10 + len(verdict_df) + 2
ws_cover.merge_cells(f"B{note_row}:H{note_row}")
c = ws_cover.cell(row=note_row, column=2,
    value=f"Composite score weights → nDCG@K: 30%  |  MRR: 25%  |  Hit@K: 20%  |  Recall@K: 15%  |  Precision@K: 10%")
c.font      = Font(name="Arial", size=9, italic=True, color="595959")
c.alignment = left()

# ----------------------------------------------------------
# SHEET 2 — SUMMARY (all methods × techniques)
# ----------------------------------------------------------
ws_sum = wb.create_sheet("📈 Summary")
freeze(ws_sum, "C2")

sum_cols = ["Rank","Method","Technique","Composite","Hit@K","MRR",
            "Precision@K","nDCG@K","Recall@K","Avg Rank","Redundancy",
            "Diversity","Boundary","N Questions"]
apply_header_row(ws_sum, 1, sum_cols)

sum_display = summary_df[[
    "rank","method","technique","composite_score",
    "hit_at_k","mrr","precision_at_k","ndcg_at_k","recall_at_k",
    "avg_rank","redundancy","diversity","boundary","n_questions"
]].copy()

for r_idx, row in enumerate(sum_display.itertuples(index=False), 2):
    fill = cell_fill(ALT_ROW) if r_idx % 2 == 0 else None
    for c_idx, val in enumerate(row, 1):
        c = ws_sum.cell(row=r_idx, column=c_idx, value=val)
        c.font      = body_font(bold=(c_idx <= 3))
        c.alignment = center()
        c.border    = thin_border()
        if fill:
            c.fill = fill

# highlight top row gold
for ci in range(1, len(sum_cols)+1):
    ws_sum.cell(row=2, column=ci).fill = cell_fill("FFF2CC")
    ws_sum.cell(row=2, column=ci).font = body_font(bold=True)

# color scale on composite (col 4)
n_rows = len(sum_display) + 1
ws_sum.conditional_formatting.add(
    f"D2:D{n_rows}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B")
)
# color scale on nDCG (col 8)
ws_sum.conditional_formatting.add(
    f"H2:H{n_rows}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B")
)

set_col_widths(ws_sum, [6,16,16,12,9,9,12,9,10,10,12,10,10,12])
ws_sum.row_dimensions[1].height = 32

# ----------------------------------------------------------
# SHEET 3 — METHOD × METRIC HEATMAP
# ----------------------------------------------------------
ws_heat = wb.create_sheet("🔥 Heatmap")

metrics   = ["hit_at_k","mrr","precision_at_k","ndcg_at_k","recall_at_k","composite_score"]
m_labels  = ["Hit@K","MRR","Precision@K","nDCG@K","Recall@K","Composite"]

# best technique per method for heatmap (hybrid_rerank usually best)
heat_data = summary_df[summary_df["technique"] == "hybrid_rerank"].copy()
if len(heat_data) == 0:
    heat_data = summary_df.groupby("method").first().reset_index()

ws_heat.merge_cells("A1:H1")
c = ws_heat["A1"]
c.value     = "Method Performance Heatmap  (Technique: hybrid_rerank)"
c.font      = Font(name="Arial", size=13, bold=True, color=DARK_BLUE)
c.alignment = center()
ws_heat.row_dimensions[1].height = 28

header_row = ["Method"] + m_labels
apply_header_row(ws_heat, 2, header_row)

for r_idx, row in enumerate(heat_data.itertuples(index=False), 3):
    ws_heat.cell(row=r_idx, column=1, value=row.method).font = body_font(bold=True)
    ws_heat.cell(row=r_idx, column=1).border = thin_border()
    ws_heat.cell(row=r_idx, column=1).alignment = left()
    for ci, met in enumerate(metrics, 2):
        val = getattr(row, met)
        c   = ws_heat.cell(row=r_idx, column=ci, value=round(val, 4))
        c.font      = body_font()
        c.alignment = center()
        c.border    = thin_border()
    ws_heat.row_dimensions[r_idx].height = 20

n_heat = len(heat_data)
for ci in range(2, len(metrics)+2):
    col_letter = get_column_letter(ci)
    ws_heat.conditional_formatting.add(
        f"{col_letter}3:{col_letter}{n_heat+2}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B")
    )

set_col_widths(ws_heat, [18,12,12,14,12,12,14])

# ----------------------------------------------------------
# SHEET 4 — PER-TECHNIQUE BREAKDOWN
# ----------------------------------------------------------
ws_tech = wb.create_sheet("🔬 By Technique")
freeze(ws_tech, "C2")

tech_cols = ["Method","Technique","Hit@K","MRR","Precision@K",
             "nDCG@K","Recall@K","Avg Rank","Redundancy","Composite"]
apply_header_row(ws_tech, 1, tech_cols)

tech_display = summary_df[[
    "method","technique","hit_at_k","mrr","precision_at_k",
    "ndcg_at_k","recall_at_k","avg_rank","redundancy","composite_score"
]].sort_values(["technique","composite_score"], ascending=[True,False])

prev_tech = None
r_idx = 2
for row in tech_display.itertuples(index=False):
    if row.technique != prev_tech:
        # technique section header
        ws_tech.merge_cells(f"A{r_idx}:J{r_idx}")
        c = ws_tech.cell(row=r_idx, column=1, value=f"  ▶  {row.technique.upper()}")
        c.font      = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill      = header_fill(MED_BLUE)
        c.alignment = left()
        ws_tech.row_dimensions[r_idx].height = 20
        r_idx += 1
        prev_tech = row.technique

    fill = cell_fill(ALT_ROW) if r_idx % 2 == 0 else None
    vals = list(row)
    for ci, val in enumerate(vals, 1):
        c = ws_tech.cell(row=r_idx, column=ci, value=val)
        c.font      = body_font()
        c.alignment = center()
        c.border    = thin_border()
        if fill:
            c.fill = fill
    r_idx += 1

set_col_widths(ws_tech, [16,16,9,9,12,9,10,10,12,12])
ws_tech.row_dimensions[1].height = 28

# ----------------------------------------------------------
# SHEET 5 — CHUNK QUALITY STATS
# ----------------------------------------------------------
ws_stats = wb.create_sheet("📦 Chunk Stats")

stats_cols = ["Method","# Chunks","Avg Words","Std Words","Min Words",
              "Max Words","Median Words","Boundary Ratio","Collection Redundancy"]
apply_header_row(ws_stats, 1, stats_cols)

for r_idx, row in enumerate(stats_df.itertuples(index=False), 2):
    fill = cell_fill(ALT_ROW) if r_idx % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws_stats.cell(row=r_idx, column=ci, value=val)
        c.font      = body_font()
        c.alignment = center()
        c.border    = thin_border()
        if fill:
            c.fill = fill

# color redundancy column red-green
n_stats = len(stats_df)
ws_stats.conditional_formatting.add(
    f"I2:I{n_stats+1}",
    ColorScaleRule(start_type="min", start_color="63BE7B",
                   end_type="max", end_color="F8696B")
)

set_col_widths(ws_stats, [16,12,12,12,12,12,14,16,20])
ws_stats.row_dimensions[1].height = 28

# ----------------------------------------------------------
# SHEET 6 — SIGNIFICANCE TESTS
# ----------------------------------------------------------
ws_sig = wb.create_sheet("🧪 Significance")

sig_cols = ["Technique","Method A","Method B","Mean A","Mean B",
            "T-stat","P-value","Significant?","Better Method"]
apply_header_row(ws_sig, 1, sig_cols)

for r_idx, row in enumerate(sig_df.itertuples(index=False), 2):
    fill = cell_fill(ALT_ROW) if r_idx % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws_sig.cell(row=r_idx, column=ci, value=val)
        c.font      = body_font()
        c.alignment = center()
        c.border    = thin_border()
        if fill:
            c.fill = fill
    # colour significance column
    sig_cell = ws_sig.cell(row=r_idx, column=8)
    if sig_cell.value == "Yes":
        sig_cell.fill = cell_fill(GREEN_GOOD)
        sig_cell.font = body_font(bold=True, color="375623")
    else:
        sig_cell.fill = cell_fill(RED_BAD)
        sig_cell.font = body_font(color="9C0006")

set_col_widths(ws_sig, [18,16,16,10,10,10,10,14,16])
ws_sig.row_dimensions[1].height = 28

# ----------------------------------------------------------
# SHEET 7 — RAW RESULTS
# ----------------------------------------------------------
ws_raw = wb.create_sheet("📋 Raw Results")
freeze(ws_raw, "D2")

raw_cols = ["Method","Technique","Q_ID","Question",
            "Hit@K","MRR","Precision@K","nDCG@K","Recall@K",
            "Avg Rank","Redundancy","Diversity","Boundary"]
apply_header_row(ws_raw, 1, raw_cols)

raw_display = raw_df[[
    "method","technique","question_id","question",
    "hit@k","mrr","precision@k","ndcg@k","recall@k",
    "avg_rank","redundancy","diversity","boundary"
]]

for r_idx, row in enumerate(raw_display.itertuples(index=False), 2):
    fill = cell_fill(ALT_ROW) if r_idx % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws_raw.cell(row=r_idx, column=ci, value=val)
        c.font      = body_font(size=9)
        c.alignment = left() if ci == 4 else center()
        c.border    = thin_border()
        if fill:
            c.fill = fill

set_col_widths(ws_raw, [14,16,6,50,8,8,12,9,10,10,12,10,10])
ws_raw.row_dimensions[1].height = 28

# ----------------------------------------------------------
# SHEET 8 — METRIC GUIDE
# ----------------------------------------------------------
ws_guide = wb.create_sheet("ℹ️ Metric Guide")

guide_data = [
    ["Metric", "Full Name", "Range", "Higher is", "Formula / Notes"],
    ["Hit@K",       "Hit Rate at K",              "0–1",    "Better",  "1 if any relevant chunk in top-K, else 0"],
    ["MRR",         "Mean Reciprocal Rank",        "0–1",    "Better",  "1/rank of first relevant chunk"],
    ["Precision@K", "Precision at K",              "0–1",    "Better",  "# relevant in top-K / K"],
    ["nDCG@K",      "Normalised Discounted CG",    "0–1",    "Better",  "DCG / IDCG; penalises late relevant hits"],
    ["Recall@K",    "Recall at K",                 "0–1",    "Better",  "# gold spans found in top-K / total gold spans"],
    ["Avg Rank",    "Average Rank of 1st Hit",     "1–K",    "Lower",   "Position of first relevant chunk (NaN if miss)"],
    ["Redundancy",  "Chunk Redundancy",            "0–1",    "Lower",   "Mean pairwise cosine sim of retrieved chunks"],
    ["Diversity",   "Chunk Diversity",             "0–1",    "Higher",  "1 − Redundancy"],
    ["Boundary",    "Sentence Boundary Ratio",     "0–1",    "Higher",  "Fraction of chunks ending with . ! ?"],
    ["Composite",   "Weighted Composite Score",    "0–1",    "Better",  "nDCG×0.30 + MRR×0.25 + Hit×0.20 + Recall×0.15 + Precision×0.10"],
    ["Coll. Redun.","Collection-level Redundancy", "0–1",    "Lower",   "Mean pairwise sim across chunk corpus sample"],
]

hdr = guide_data[0]
apply_header_row(ws_guide, 1, hdr)
for r_idx, row in enumerate(guide_data[1:], 2):
    fill = cell_fill(ALT_ROW) if r_idx % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws_guide.cell(row=r_idx, column=ci, value=val)
        c.font      = body_font()
        c.alignment = left() if ci in [2,5] else center()
        c.border    = thin_border()
        if fill:
            c.fill = fill
        if ci == 4:
            c.font = body_font(bold=True,
                color="375623" if val == "Better" else ("9C0006" if val == "Lower" else "000000"))

set_col_widths(ws_guide, [14,30,10,12,52])
ws_guide.row_dimensions[1].height = 28

# ----------------------------------------------------------
# SAVE
# ----------------------------------------------------------
OUTPUT_XLSX = "benchmark_report.xlsx"
wb.save(OUTPUT_XLSX)
print(f"\nExcel report saved: {OUTPUT_XLSX}")
print("\n✅  All outputs:")
print("   raw_results.csv       — per-question scores")
print("   chunk_stats.csv       — chunk quality per method")
print("   summary.csv           — aggregated scores ranked")
print("   benchmark_report.xlsx — full formatted report\n")