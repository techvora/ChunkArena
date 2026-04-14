"""
RAG Chunking Strategy Benchmark — Production Grade (Standard Redundancy)
========================================================================
Evaluates N chunking strategies × M retrieval techniques across all IR metrics.

Uses standard definitions:
- Redundancy = mean pairwise cosine similarity among top‑K chunks
- Diversity = 1 - Redundancy
- All other metrics follow classic IR formulas.

Outputs
-------
  raw_results.csv         per-question × method × technique scores
  chunk_stats.csv         chunk quality stats per method
  summary.csv             aggregated + ranked composite scores
  benchmark_report.xlsx   full formatted Excel workbook (8 sheets)
"""

import re
import warnings
import numpy as np
import pandas as pd
from collections import defaultdict
from scipy import stats
from tqdm import tqdm

from sentence_transformers import SentenceTransformer, CrossEncoder
from qdrant_client import QdrantClient
from sklearn.metrics.pairwise import cosine_similarity
from rank_bm25 import BM25Okapi
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings("ignore")

# ============================================================
# CONFIG — tune these per dataset/domain
# ============================================================
CSV_PATH        = "/home/root473/Documents/POC/ChunkArena/Golden_dataset/Banking_system.csv"
FINAL_K         = 5            # top-K results returned to user
RETRIEVAL_K     = 50           # dense/hybrid candidates before reranking
HYBRID_CAND_K   = 100          # BM25 + dense candidates for hybrid fusion

# Semantic relevance threshold (BGE-M3 on banking text)
SOFT_THRESHOLD  = 0.72

CHUNK_METHODS = [
    "fixed_size", "overlapping", "sentence",
    "paragraph", "recursive", "header", "semantic"
]
TECHNIQUES = ["dense", "hybrid", "dense_rerank", "hybrid_rerank"]

# Composite score weights (must sum to 1.0)
COMPOSITE_WEIGHTS = {
    "ndcg_at_k"     : 0.30,
    "mrr"           : 0.25,
    "hit_at_k"      : 0.20,
    "recall_at_k"   : 0.15,
    "precision_at_k": 0.10,
}

# Threshold verdicts — used for per-metric diagnostic column in Excel
THRESHOLDS = {
    "hit_at_k"      : {"good": 0.8,  "moderate": 0.5},
    "mrr"           : {"good": 0.8,  "moderate": 0.5},
    "precision_at_k": {"good": 0.7,  "moderate": 0.3},
    "ndcg_at_k"     : {"good": 0.7,  "moderate": 0.5},
    "recall_at_k"   : {"good": 0.8,  "moderate": 0.6},
    "redundancy"    : {"good": 0.3,  "moderate": 0.6},   # lower is better (mean similarity)
    "composite_score": {"good": 0.75, "moderate": 0.65},
}

# ============================================================
# LOAD GOLD DATA
# ============================================================
print("Loading gold dataset...")
gold_df = pd.read_csv(CSV_PATH)
questions_data = []
for _, row in gold_df.iterrows():
    questions_data.append({
        "id"        : len(questions_data),
        "question"  : row["Question"],
        "gold_spans": [s.strip() for s in str(row["Gold_spans"]).split(";")]
    })
print(f"  Loaded {len(questions_data)} questions")

# ============================================================
# INIT MODELS
# ============================================================
print("Loading models...")
device       = "cpu"
embedder     = SentenceTransformer("BAAI/bge-m3", device=device)
cross_encoder = CrossEncoder("cross-encoder/ms-marco-MiniLM-L-6-v2", device=device)
client       = QdrantClient(host="localhost", port=6333)
print("  Models ready")

# ============================================================
# CACHES
# ============================================================
embedding_cache  = {}
retrieval_cache  = {}
rerank_cache     = {}
relevance_cache  = {}

def get_embedding(text: str) -> np.ndarray:
    if text not in embedding_cache:
        embedding_cache[text] = embedder.encode([text], convert_to_numpy=True)[0]
    return embedding_cache[text]

# ============================================================
# LOAD CHUNKS + BM25
# ============================================================
word_re = re.compile(r"\w+")

def get_all_chunks(collection_name: str):
    all_points, offset = [], None
    while True:
        batch, next_offset = client.scroll(
            collection_name=collection_name,
            limit=1000,
            offset=offset,
            with_payload=True,
        )
        all_points.extend(batch)
        if next_offset is None:
            break
        offset = next_offset
    ids   = [p.id for p in all_points]
    texts = [p.payload["text"] for p in all_points]
    return ids, texts

all_chunks_cache     = {}
all_chunks_text_dict = {}
bm25_models          = {}

print("Loading collections and building BM25 indexes...")
for method in CHUNK_METHODS:
    ids, texts = get_all_chunks(method)
    all_chunks_cache[method]     = (ids, texts)
    all_chunks_text_dict[method] = dict(zip(ids, texts))
    tokenized = [word_re.findall(t.lower()) for t in texts]
    bm25_models[method] = BM25Okapi(tokenized)
    print(f"  {method}: {len(texts)} chunks")

# ============================================================
# CHUNK QUALITY STATS (independent of retrieval)
# ============================================================
def collection_redundancy(method: str, sample_size: int = 300) -> float:
    """
    Standard corpus-level redundancy: mean pairwise cosine similarity
    over a random sample of chunks.
    """
    _, texts = all_chunks_cache[method]
    sample   = texts[:sample_size]
    embs     = np.array([get_embedding(t) for t in sample])
    sim      = cosine_similarity(embs)
    n        = len(sim)
    triu = np.triu_indices(n, k=1)
    if n < 2:
        return 0.0
    sum_sim = np.sum(sim[triu])
    num_pairs = n * (n - 1) / 2
    return round(sum_sim / num_pairs, 4)

def chunk_stats(method: str) -> dict:
    _, texts  = all_chunks_cache[method]
    lengths   = [len(t.split()) for t in texts]
    boundary  = sum(1 for t in texts if re.search(r"[.!?][\"']?\s*$", t.strip()))
    return {
        "method"               : method,
        "num_chunks"           : len(texts),
        "avg_words"            : round(np.mean(lengths), 2),
        "std_words"            : round(np.std(lengths), 2),
        "min_words"            : int(np.min(lengths)),
        "max_words"            : int(np.max(lengths)),
        "median_words"         : round(float(np.median(lengths)), 2),
        "boundary_ratio"       : round(boundary / len(texts), 4),
        "collection_redundancy": collection_redundancy(method),
    }

# ============================================================
# RETRIEVAL
# ============================================================
def dense_search(query_text: str, query_emb: np.ndarray,
                 method: str, top_k: int):
    key = ("dense", method, query_text, top_k)
    if key in retrieval_cache:
        return retrieval_cache[key]
    res   = client.query_points(collection_name=method,
                                query=query_emb.tolist(), limit=top_k)
    ids   = [r.id for r in res.points]
    texts = [all_chunks_text_dict[method][i] for i in ids]
    retrieval_cache[key] = (ids, texts)
    return ids, texts

def hybrid_search(query: str, query_emb: np.ndarray,
                  method: str, top_k: int):
    key = ("hybrid", method, query, top_k)
    if key in retrieval_cache:
        return retrieval_cache[key]

    dense_ids, _ = dense_search(query, query_emb, method, HYBRID_CAND_K)

    bm25    = bm25_models[method]
    tokens  = word_re.findall(query.lower())
    scores  = bm25.get_scores(tokens)
    all_ids, _ = all_chunks_cache[method]
    bm25_ranked = sorted(zip(all_ids, scores),
                         key=lambda x: x[1], reverse=True)[:HYBRID_CAND_K]

    # Reciprocal Rank Fusion (standard RRF)
    fused = defaultdict(float)
    for i, cid in enumerate(dense_ids):
        fused[cid] += 1 / (60 + i)
    for i, (cid, _) in enumerate(bm25_ranked):
        fused[cid] += 1 / (60 + i)

    final_ids = sorted(fused, key=lambda x: fused[x], reverse=True)[:top_k]
    texts     = [all_chunks_text_dict[method][i] for i in final_ids]
    retrieval_cache[key] = (final_ids, texts)
    return final_ids, texts

def rerank(query: str, ids: list, texts: list,
           top_k: int, tag: str = "") -> tuple:
    """Rerank with cross-encoder; cache key includes tag to separate techniques."""
    if not texts:
        return [], []
    key = (query, tag, tuple(ids[:20]))
    if key in rerank_cache:
        s_ids, s_texts = rerank_cache[key]
        return s_ids[:top_k], s_texts[:top_k]

    pairs  = [[query, t] for t in texts]
    scores = cross_encoder.predict(pairs)
    idx    = np.argsort(scores)[::-1]
    s_ids  = [ids[i] for i in idx]
    s_texts = [texts[i] for i in idx]
    rerank_cache[key] = (s_ids, s_texts)
    return s_ids[:top_k], s_texts[:top_k]

# ============================================================
# RELEVANCE — exact match + semantic fallback
# ============================================================
def is_relevant(chunk: str, gold_spans: list) -> bool:
    """Chunk relevant if exact match or cosine similarity >= SOFT_THRESHOLD."""
    key = (chunk, tuple(gold_spans))
    if key in relevance_cache:
        return relevance_cache[key]

    chunk_lower = chunk.lower()
    for span in gold_spans:
        if span.lower() in chunk_lower:
            relevance_cache[key] = True
            return True

    chunk_emb = get_embedding(chunk)
    for span in gold_spans:
        score = cosine_similarity([chunk_emb], [get_embedding(span)])[0][0]
        if score >= SOFT_THRESHOLD:
            relevance_cache[key] = True
            return True

    relevance_cache[key] = False
    return False

# ============================================================
# METRIC FUNCTIONS
# ============================================================
def hit_at_k(chunks: list, gold_spans: list, k: int) -> int:
    return int(any(is_relevant(c, gold_spans) for c in chunks[:k]))

def mrr_score(chunks: list, gold_spans: list) -> float:
    for i, c in enumerate(chunks, 1):
        if is_relevant(c, gold_spans):
            return 1.0 / i
    return 0.0

def precision_at_k(chunks: list, gold_spans: list, k: int) -> float:
    if k == 0:
        return 0.0
    return sum(1 for c in chunks[:k] if is_relevant(c, gold_spans)) / k

def ndcg_at_k(chunks: list, gold_spans: list, k: int) -> float:
    dcg = sum(1 / np.log2(i + 2)
              for i, c in enumerate(chunks[:k])
              if is_relevant(c, gold_spans))
    idcg = sum(1 / np.log2(i + 2)
               for i in range(min(len(gold_spans), k)))
    return dcg / idcg if idcg > 0 else 0.0

def recall_at_k(chunks: list, gold_spans: list, k: int) -> float:
    if not gold_spans:
        return 0.0
    found = sum(1 for span in gold_spans
                if any(is_relevant(c, [span]) for c in chunks[:k]))
    return found / len(gold_spans)

def avg_rank_score(chunks: list, gold_spans: list) -> float:
    for i, c in enumerate(chunks, 1):
        if is_relevant(c, gold_spans):
            return float(i)
    return np.nan

def redundancy_score(chunks: list) -> float:
    """Standard: mean pairwise cosine similarity among top-K chunks."""
    if len(chunks) < 2:
        return 0.0
    embs = np.array([get_embedding(c) for c in chunks])
    sim = cosine_similarity(embs)
    n = len(sim)
    triu_indices = np.triu_indices(n, k=1)
    sum_sim = np.sum(sim[triu_indices])
    num_pairs = n * (n - 1) / 2
    return round(sum_sim / num_pairs, 4)

def boundary_score(chunks: list) -> float:
    if not chunks:
        return 0.0
    return round(sum(1 for c in chunks if re.search(r"[.!?][\"']?\s*$", c.strip())) / len(chunks), 4)

def threshold_verdict(metric: str, value: float) -> str:
    if metric not in THRESHOLDS or pd.isna(value):
        return "N/A"
    t = THRESHOLDS[metric]
    if metric == "redundancy":   # lower is better
        if value <= t["good"]:
            return "Good"
        elif value <= t["moderate"]:
            return "Moderate"
        else:
            return "Bad"
    else:                         # higher is better
        if value >= t["good"]:
            return "Good"
        elif value >= t["moderate"]:
            return "Moderate"
        else:
            return "Bad"

# ============================================================
# CHUNK QUALITY STATS
# ============================================================
print("\nComputing chunk quality stats...")
stats_rows = []
for method in CHUNK_METHODS:
    print(f"  {method}...")
    stats_rows.append(chunk_stats(method))
stats_df = pd.DataFrame(stats_rows)

# ============================================================
# MAIN EVALUATION LOOP
# ============================================================
print("\nRunning evaluation...")
raw_results = []

for method in CHUNK_METHODS:
    print(f"\n  Method: {method}")
    for q in tqdm(questions_data, desc=f"    {method}"):
        query      = q["question"]
        gold_spans = q["gold_spans"]
        q_emb      = get_embedding(query)

        dense_ids,  dense_texts  = dense_search(query, q_emb, method, RETRIEVAL_K)
        hybrid_ids, hybrid_texts = hybrid_search(query, q_emb, method, RETRIEVAL_K)

        techniques = {
            "dense"         : (dense_ids[:FINAL_K],  dense_texts[:FINAL_K]),
            "hybrid"        : (hybrid_ids[:FINAL_K], hybrid_texts[:FINAL_K]),
            "dense_rerank"  : rerank(query, dense_ids,  dense_texts,  FINAL_K, tag="dense"),
            "hybrid_rerank" : rerank(query, hybrid_ids, hybrid_texts, FINAL_K, tag="hybrid"),
        }

        for tech, (ids, texts) in techniques.items():
            h   = hit_at_k(texts, gold_spans, FINAL_K)
            m   = mrr_score(texts, gold_spans)
            p   = precision_at_k(texts, gold_spans, FINAL_K)
            n   = ndcg_at_k(texts, gold_spans, FINAL_K)
            rc  = recall_at_k(texts, gold_spans, FINAL_K)
            ar  = avg_rank_score(texts, gold_spans)
            rd  = redundancy_score(texts)
            b   = boundary_score(texts)
            miss = 1 if pd.isna(ar) else 0

            raw_results.append({
                "method"       : method,
                "technique"    : tech,
                "question_id"  : q["id"],
                "question"     : query,
                "hit@k"        : h,
                "mrr"          : round(m,  4),
                "precision@k"  : round(p,  4),
                "ndcg@k"       : round(n,  4),
                "recall@k"     : round(rc, 4),
                "avg_rank"     : ar,
                "miss"         : miss,
                "redundancy"   : rd,
                "diversity"    : round(1 - rd, 4),
                "boundary"     : b,
            })

raw_df = pd.DataFrame(raw_results)
raw_df.to_csv("raw_results.csv", index=False)
stats_df.to_csv("chunk_stats.csv", index=False)
print("\nRaw results saved.")

# ============================================================
# SUMMARY — aggregate per (method, technique)
# ============================================================
summary_df = raw_df.groupby(["method", "technique"]).agg(
    hit_at_k        = ("hit@k",       "mean"),
    mrr             = ("mrr",         "mean"),
    precision_at_k  = ("precision@k", "mean"),
    ndcg_at_k       = ("ndcg@k",      "mean"),
    recall_at_k     = ("recall@k",    "mean"),
    avg_rank        = ("avg_rank",    "mean"),
    miss_rate       = ("miss",        "mean"),
    redundancy      = ("redundancy",  "mean"),
    diversity       = ("diversity",   "mean"),
    boundary        = ("boundary",    "mean"),
    n_questions     = ("question_id", "count"),
).round(4).reset_index()

# Composite score
summary_df["composite_score"] = sum(
    summary_df[m] * w for m, w in COMPOSITE_WEIGHTS.items()
).round(4)

summary_df = summary_df.sort_values(
    "composite_score", ascending=False
).reset_index(drop=True)
summary_df["rank"] = summary_df.index + 1

# Threshold verdicts
for metric in ["hit_at_k", "mrr", "ndcg_at_k", "recall_at_k",
               "redundancy", "composite_score"]:
    summary_df[f"{metric}_verdict"] = summary_df[metric].apply(
        lambda v: threshold_verdict(metric, v)
    )

summary_df.to_csv("summary.csv", index=False)

# ============================================================
# SIGNIFICANCE TESTS
# ============================================================
sig_rows = []
for tech in TECHNIQUES:
    sub = raw_df[raw_df["technique"] == tech]
    for i, m1 in enumerate(CHUNK_METHODS):
        for m2 in CHUNK_METHODS[i + 1:]:
            s1 = sub[sub["method"] == m1]["ndcg@k"].dropna()
            s2 = sub[sub["method"] == m2]["ndcg@k"].dropna()
            if len(s1) > 1 and len(s2) > 1:
                t_stat, p_val = stats.ttest_rel(s1, s2)
                sig_rows.append({
                    "technique"  : tech,
                    "method_a"   : m1,
                    "method_b"   : m2,
                    "mean_a"     : round(s1.mean(), 4),
                    "mean_b"     : round(s2.mean(), 4),
                    "t_stat"     : round(t_stat, 4),
                    "p_value"    : round(p_val,  4),
                    "significant": "Yes" if p_val < 0.05 else "No",
                    "better"     : m1 if s1.mean() > s2.mean() else m2,
                })
sig_df = pd.DataFrame(sig_rows)

# ============================================================
# VERDICT — best per technique & overall winner
# ============================================================
verdict_rows = []
for tech in TECHNIQUES:
    sub  = summary_df[summary_df["technique"] == tech]
    best = sub.sort_values("composite_score", ascending=False).iloc[0]
    verdict_rows.append({
        "technique"      : tech,
        "best_method"    : best["method"],
        "composite_score": best["composite_score"],
        "hit@k"          : best["hit_at_k"],
        "mrr"            : best["mrr"],
        "ndcg@k"         : best["ndcg_at_k"],
        "recall@k"       : best["recall_at_k"],
        "precision@k"    : best["precision_at_k"],
        "miss_rate"      : best["miss_rate"],
        "redundancy"     : best["redundancy"],
    })
verdict_df = pd.DataFrame(verdict_rows)

overall_best = summary_df.iloc[0]
print(f"\n{'='*60}")
print(f"OVERALL BEST: {overall_best['method']} + {overall_best['technique']}")
print(f"  Composite   : {overall_best['composite_score']}")
print(f"  nDCG@{FINAL_K}      : {overall_best['ndcg_at_k']}")
print(f"  MRR         : {overall_best['mrr']}")
print(f"  Hit@{FINAL_K}       : {overall_best['hit_at_k']}")
print(f"  Miss rate   : {overall_best['miss_rate']:.1%}")
print(f"  Redundancy  : {overall_best['redundancy']} (mean pairwise similarity)")
print(f"{'='*60}")

# ============================================================
# EXCEL WORKBOOK (8 sheets)
# ============================================================
print("\nBuilding Excel report...")

DARK_BLUE  = "1F3864"
MED_BLUE   = "2F5496"
HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
ALT_ROW    = "EBF3FB"
GREEN_GOOD = "C6EFCE"
GREEN_FG   = "375623"
RED_BAD    = "FFC7CE"
RED_FG     = "9C0006"
AMBER_MID  = "FFEB9C"
AMBER_FG   = "7D6608"
WHITE      = "FFFFFF"

def hfont(size=11, bold=True, color=HEADER_FG):
    return Font(name="Arial", size=size, bold=bold, color=color)

def bfont(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def hfill(color=HEADER_BG):
    return PatternFill("solid", fgColor=color)

def cfill(color):
    return PatternFill("solid", fgColor=color)

def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def ca():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def la():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hrow(ws, row_num, cols):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font = hfont(); c.fill = hfill()
        c.alignment = ca(); c.border = tborder()

def drows(ws, dataframe, start_row=2):
    for r_idx, row in enumerate(dataframe.itertuples(index=False), start_row):
        alt = cfill(ALT_ROW) if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = bfont(); c.alignment = ca(); c.border = tborder()
            if alt: c.fill = alt

def col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def verdict_fill(v):
    if v == "Good":     return cfill(GREEN_GOOD), bfont(bold=True, color=GREEN_FG)
    if v == "Moderate": return cfill(AMBER_MID),  bfont(bold=True, color=AMBER_FG)
    if v == "Bad":      return cfill(RED_BAD),     bfont(bold=True, color=RED_FG)
    return None, bfont()

wb = Workbook()
wb.remove(wb.active)

# SHEET 1 — Verdict / Cover
ws_v = wb.create_sheet("📊 Verdict")
ws_v.column_dimensions["A"].width = 3
for col in "BCDEFGHIJ":
    ws_v.column_dimensions[col].width = 18

ws_v.merge_cells("B2:J2")
c = ws_v["B2"]
c.value = "RAG Chunking Strategy Benchmark — Final Verdict"
c.font = Font(name="Arial", size=18, bold=True, color=DARK_BLUE)
c.alignment = ca()
ws_v.row_dimensions[2].height = 36

ws_v.merge_cells("B3:J3")
c = ws_v["B3"]
c.value = (f"Evaluated {len(CHUNK_METHODS)} methods × {len(TECHNIQUES)} techniques | "
           f"{len(questions_data)} questions | K={FINAL_K} | "
           f"Soft threshold={SOFT_THRESHOLD} | Redundancy = mean pairwise similarity")
c.font = Font(name="Arial", size=10, italic=True, color="595959")
c.alignment = ca()
ws_v.row_dimensions[3].height = 18

ws_v.merge_cells("B5:J7")
c = ws_v["B5"]
c.value = (f"🏆  OVERALL WINNER:  {overall_best['method'].upper()}  +  "
           f"{overall_best['technique'].upper()}   "
           f"(Composite: {overall_best['composite_score']:.4f})")
c.font = Font(name="Arial", size=14, bold=True, color=WHITE)
c.fill = hfill(MED_BLUE); c.alignment = ca()
for r in [5,6,7]:
    ws_v.row_dimensions[r].height = 22

hdrs = ["Technique", "Best Method", "Composite", "Hit@K", "MRR",
        "nDCG@K", "Recall@K", "Precision@K", "Miss Rate", "Redundancy"]
hrow(ws_v, 9, hdrs)
ws_v.row_dimensions[9].height = 22

for r, row in verdict_df.iterrows():
    rn = 10 + r
    vals = [row["technique"], row["best_method"], row["composite_score"],
            row["hit@k"], row["mrr"], row["ndcg@k"],
            row["recall@k"], row["precision@k"],
            f"{row['miss_rate']:.1%}", row["redundancy"]]
    for ci, val in enumerate(vals, 2):
        c = ws_v.cell(row=rn, column=ci, value=val)
        c.font = bfont(bold=(ci <= 3))
        c.alignment = ca(); c.border = tborder()
        if ci in [3,4]:
            c.fill = cfill(GREEN_GOOD)
    ws_v.row_dimensions[rn].height = 20

note_row = 10 + len(verdict_df) + 2
ws_v.merge_cells(f"B{note_row}:J{note_row}")
c = ws_v.cell(row=note_row, column=2,
    value=(f"Composite weights → nDCG@K:{COMPOSITE_WEIGHTS['ndcg_at_k']*100:.0f}%  "
           f"MRR:{COMPOSITE_WEIGHTS['mrr']*100:.0f}%  "
           f"Hit@K:{COMPOSITE_WEIGHTS['hit_at_k']*100:.0f}%  "
           f"Recall@K:{COMPOSITE_WEIGHTS['recall_at_k']*100:.0f}%  "
           f"Precision@K:{COMPOSITE_WEIGHTS['precision_at_k']*100:.0f}%"))
c.font = Font(name="Arial", size=9, italic=True, color="595959")
c.alignment = la()

# SHEET 2 — Summary
ws_s = wb.create_sheet("📈 Summary")
ws_s.freeze_panes = "C2"
sum_cols = ["Rank", "Method", "Technique", "Composite", "Verdict",
            "Hit@K", "MRR", "Precision@K", "nDCG@K", "Recall@K",
            "Avg Rank", "Miss Rate", "Redundancy", "Diversity",
            "Boundary", "N Questions"]
hrow(ws_s, 1, sum_cols)

sum_disp = summary_df[[
    "rank", "method", "technique", "composite_score", "composite_score_verdict",
    "hit_at_k", "mrr", "precision_at_k", "ndcg_at_k", "recall_at_k",
    "avg_rank", "miss_rate", "redundancy", "diversity", "boundary", "n_questions"
]].copy()

for r_idx, row in enumerate(sum_disp.itertuples(index=False), 2):
    alt = cfill(ALT_ROW) if r_idx % 2 == 0 else None
    vals = list(row)
    for c_idx, val in enumerate(vals, 1):
        c = ws_s.cell(row=r_idx, column=c_idx, value=val)
        c.font = bfont(bold=(c_idx <= 3))
        c.alignment = ca(); c.border = tborder()
        if alt: c.fill = alt
    verdict_val = vals[4]
    fill, font = verdict_fill(verdict_val)
    if fill:
        ws_s.cell(row=r_idx, column=5).fill = fill
        ws_s.cell(row=r_idx, column=5).font = font

for ci in range(1, len(sum_cols)+1):
    ws_s.cell(row=2, column=ci).fill = cfill("FFF2CC")
    ws_s.cell(row=2, column=ci).font = bfont(bold=True)

n_rows = len(sum_disp) + 1
for col_letter in ["D", "I"]:
    ws_s.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{n_rows}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B")
    )
ws_s.conditional_formatting.add(
    f"M2:M{n_rows}",
    ColorScaleRule(start_type="min", start_color="63BE7B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="F8696B")
)
col_widths(ws_s, [6,16,16,12,12,9,9,12,9,10,10,10,12,10,10,12])
ws_s.row_dimensions[1].height = 32

# SHEET 3 — Heatmap (hybrid_rerank)
ws_h = wb.create_sheet("🔥 Heatmap")
metrics_heat = ["hit_at_k", "mrr", "precision_at_k", "ndcg_at_k",
                "recall_at_k", "redundancy", "composite_score"]
m_labels_heat = ["Hit@K", "MRR", "Precision@K", "nDCG@K",
                 "Recall@K", "Redundancy", "Composite"]
heat_data = summary_df[summary_df["technique"] == "hybrid_rerank"].copy()
if len(heat_data) == 0:
    heat_data = summary_df.groupby("method").first().reset_index()
ws_h.merge_cells("A1:H1")
c = ws_h["A1"]
c.value = "Performance heatmap (technique: hybrid_rerank)"
c.font = Font(name="Arial", size=13, bold=True, color=DARK_BLUE)
c.alignment = ca()
ws_h.row_dimensions[1].height = 28
hrow(ws_h, 2, ["Method"] + m_labels_heat)
for r_idx, row in enumerate(heat_data.itertuples(index=False), 3):
    c = ws_h.cell(row=r_idx, column=1, value=row.method)
    c.font = bfont(bold=True); c.border = tborder(); c.alignment = la()
    for ci, met in enumerate(metrics_heat, 2):
        val = getattr(row, met)
        cell = ws_h.cell(row=r_idx, column=ci, value=round(val,4))
        cell.font = bfont(); cell.alignment = ca(); cell.border = tborder()
    ws_h.row_dimensions[r_idx].height = 20
n_heat = len(heat_data)
for ci in range(2, len(metrics_heat)+1):
    if ci == 7: continue
    cl = get_column_letter(ci)
    ws_h.conditional_formatting.add(
        f"{cl}3:{cl}{n_heat+2}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B")
    )
ws_h.conditional_formatting.add(
    f"G3:G{n_heat+2}",
    ColorScaleRule(start_type="min", start_color="63BE7B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="F8696B")
)
col_widths(ws_h, [18,12,12,14,12,12,14,14])

# SHEET 4 — By Technique
ws_t = wb.create_sheet("🔬 By Technique")
ws_t.freeze_panes = "C2"
tech_cols = ["Method", "Technique", "Hit@K", "MRR", "Precision@K",
             "nDCG@K", "Recall@K", "Avg Rank", "Miss Rate",
             "Redundancy", "Composite", "Verdict"]
hrow(ws_t, 1, tech_cols)
tech_disp = summary_df[[
    "method", "technique", "hit_at_k", "mrr", "precision_at_k",
    "ndcg_at_k", "recall_at_k", "avg_rank", "miss_rate",
    "redundancy", "composite_score", "composite_score_verdict"
]].sort_values(["technique", "composite_score"], ascending=[True, False])
prev_tech = None
r_idx = 2
for row in tech_disp.itertuples(index=False):
    if row.technique != prev_tech:
        ws_t.merge_cells(f"A{r_idx}:L{r_idx}")
        c = ws_t.cell(row=r_idx, column=1, value=f"  ▶  {row.technique.upper()}")
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = hfill(MED_BLUE); c.alignment = la()
        ws_t.row_dimensions[r_idx].height = 20
        r_idx += 1
        prev_tech = row.technique
    alt = cfill(ALT_ROW) if r_idx % 2 == 0 else None
    vals = list(row)
    for ci, val in enumerate(vals, 1):
        c = ws_t.cell(row=r_idx, column=ci, value=val)
        c.font = bfont(); c.alignment = ca(); c.border = tborder()
        if alt: c.fill = alt
    vfill, vfont = verdict_fill(vals[11])
    if vfill:
        ws_t.cell(row=r_idx, column=12).fill = vfill
        ws_t.cell(row=r_idx, column=12).font = vfont
    r_idx += 1
col_widths(ws_t, [16,16,9,9,12,9,10,10,10,12,12,10])
ws_t.row_dimensions[1].height = 28

# SHEET 5 — Chunk Stats
ws_cs = wb.create_sheet("📦 Chunk Stats")
stats_cols = ["Method", "# Chunks", "Avg Words", "Std Words", "Min Words",
              "Max Words", "Median Words", "Boundary Ratio", "Coll. Redundancy"]
hrow(ws_cs, 1, stats_cols)
drows(ws_cs, stats_df)
n_stats = len(stats_df)
ws_cs.conditional_formatting.add(
    f"I2:I{n_stats+1}",
    ColorScaleRule(start_type="min", start_color="63BE7B",
                   end_type="max", end_color="F8696B")
)
col_widths(ws_cs, [16,12,12,12,12,12,14,16,20])
ws_cs.row_dimensions[1].height = 28

# SHEET 6 — Significance
ws_sig = wb.create_sheet("🧪 Significance")
sig_cols = ["Technique", "Method A", "Method B", "Mean A", "Mean B",
            "T-stat", "P-value", "Significant?", "Better Method"]
hrow(ws_sig, 1, sig_cols)
for r_idx, row in enumerate(sig_df.itertuples(index=False), 2):
    alt = cfill(ALT_ROW) if r_idx % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws_sig.cell(row=r_idx, column=ci, value=val)
        c.font = bfont(); c.alignment = ca(); c.border = tborder()
        if alt: c.fill = alt
    sig_cell = ws_sig.cell(row=r_idx, column=8)
    if sig_cell.value == "Yes":
        sig_cell.fill = cfill(GREEN_GOOD)
        sig_cell.font = bfont(bold=True, color=GREEN_FG)
    else:
        sig_cell.fill = cfill(RED_BAD)
        sig_cell.font = bfont(color=RED_FG)
col_widths(ws_sig, [18,16,16,10,10,10,10,14,16])
ws_sig.row_dimensions[1].height = 28

# SHEET 7 — Raw Results
ws_r = wb.create_sheet("📋 Raw Results")
ws_r.freeze_panes = "E2"
raw_cols = ["Method", "Technique", "Q_ID", "Question",
            "Hit@K", "MRR", "Precision@K", "nDCG@K", "Recall@K",
            "Avg Rank", "Miss", "Redundancy", "Diversity", "Boundary"]
hrow(ws_r, 1, raw_cols)
raw_disp = raw_df[[
    "method", "technique", "question_id", "question",
    "hit@k", "mrr", "precision@k", "ndcg@k", "recall@k",
    "avg_rank", "miss", "redundancy", "diversity", "boundary"
]]
for r_idx, row in enumerate(raw_disp.itertuples(index=False), 2):
    alt = cfill(ALT_ROW) if r_idx % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws_r.cell(row=r_idx, column=ci, value=val)
        c.font = bfont(size=9)
        c.alignment = la() if ci == 4 else ca()
        c.border = tborder()
        if alt: c.fill = alt
col_widths(ws_r, [14,16,6,50,8,8,12,9,10,10,7,12,10,10])
ws_r.row_dimensions[1].height = 28

# SHEET 8 — Metric Guide
ws_g = wb.create_sheet("ℹ️ Metric Guide")
guide = [
    ["Metric", "Definition", "Range", "Higher is", "Threshold: Good", "Threshold: Moderate", "Threshold: Bad", "Formula / Notes"],
    ["Hit@K",        "≥1 relevant chunk in top-K",         "0–1",   "Better", "≥ 0.80", "0.50–0.80", "< 0.50", "Binary per question, averaged"],
    ["MRR",          "Rank of first relevant chunk",        "~0–1",  "Better", "≥ 0.80", "0.50–0.80", "< 0.50", "1/rank of first hit; 0 on miss"],
    ["Precision@K",  "Relevant fraction of top-K",          "0–1",   "Better", "≥ 0.70", "0.30–0.70", "< 0.30", "Relevant in K / K"],
    ["nDCG@K",       "Ranking quality (position-penalised)","0–1",   "Better", "≥ 0.70", "0.50–0.70", "< 0.50", "DCG / IDCG; penalises late hits"],
    ["Recall@K",     "Gold spans covered in top-K",         "0–1",   "Better", "≥ 0.80", "0.60–0.80", "< 0.60", "Spans found / total spans"],
    ["Avg Rank",     "Position of first hit",               "1–K",   "Lower",  "1–2",    "2–3",       "> 3",    "NaN on miss; miss_rate tracked"],
    ["Redundancy",   "Mean pairwise cosine similarity among top-K chunks", "0–1", "Lower", "< 0.3", "0.3–0.6", "> 0.6", "Average similarity (0=diverse, 1=redundant)"],
    ["Diversity",    "1 − Redundancy",                      "0–1",   "Higher", "> 0.6", "0.3–0.6", "< 0.3", "Inverse of redundancy"],
    ["Boundary",     "Sentence-boundary chunk endings",     "0–1",   "Higher", "Context", "-",        "-",       "Fraction ending in . ! ?"],
    ["Composite",    "Weighted combination of key metrics", "0–1",   "Better", "≥ 0.75", "0.65–0.75", "< 0.65",
     f"nDCG×{COMPOSITE_WEIGHTS['ndcg_at_k']} + MRR×{COMPOSITE_WEIGHTS['mrr']} + "
     f"Hit×{COMPOSITE_WEIGHTS['hit_at_k']} + Recall×{COMPOSITE_WEIGHTS['recall_at_k']} + "
     f"Precision×{COMPOSITE_WEIGHTS['precision_at_k']}"],
    ["Coll. Redun.", "Mean pairwise cosine similarity in corpus sample", "0–1", "Lower", "< 0.3", "0.3–0.6", "> 0.6", "Average similarity (0=diverse, 1=redundant)"],
    ["Miss Rate",    "Fraction of questions with no hit",   "0–1",   "Lower",  "< 0.05", "0.05–0.20", "> 0.20", "1 − avg(hit@k) across questions"],
]
hrow(ws_g, 1, guide[0])
for r_idx, row in enumerate(guide[1:], 2):
    alt = cfill(ALT_ROW) if r_idx % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws_g.cell(row=r_idx, column=ci, value=val)
        c.font = bfont()
        c.alignment = la() if ci in [2,8] else ca()
        c.border = tborder()
        if alt: c.fill = alt
        if ci == 4:
            is_lower = (val == "Lower")
            c.font = bfont(bold=True, color=RED_FG if is_lower else GREEN_FG)
col_widths(ws_g, [14,34,8,10,14,16,14,58])
ws_g.row_dimensions[1].height = 28

# SAVE
OUTPUT_XLSX = "benchmark_report.xlsx"
wb.save(OUTPUT_XLSX)
print(f"\nExcel report saved → {OUTPUT_XLSX}")
print("\n✅  All outputs:")
print("   raw_results.csv       — per-question scores (with miss column)")
print("   chunk_stats.csv       — chunk quality per method")
print("   summary.csv           — aggregated scores + verdicts ranked")
print("   benchmark_report.xlsx — 8-sheet formatted report\n")
print("Key design choices:")
print(f"  • SOFT_THRESHOLD      : {SOFT_THRESHOLD} (semantic relevance)")
print(f"  • Redundancy          : mean pairwise cosine similarity (standard)")
print("  • Rerank cache key    : includes technique tag (no collision)")
print("  • miss_rate           : explicit column in raw + summary")
print("  • Threshold verdicts  : Good/Moderate/Bad per metric in Excel")