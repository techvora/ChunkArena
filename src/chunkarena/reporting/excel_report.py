"""Excel report builder.

Writes a single consolidated 3-sheet benchmark workbook:

    Experiment Matrix  Flat result table with one row per (method,
                       technique) including chunk quality stats,
                       retrieval metrics, and auto-generated notes.
    Raw Results        Per-question per-technique scores for drilldown.
    Heatmap            All techniques x all methods sorted by composite
                       score (high to low) with color scales for visual
                       comparison.

The top-level function build_workbook takes the summary, raw, and stats
dataframes produced by the runner and saves the styled report to
BENCHMARK_XLSX. This is the sole output artifact — no separate CSV files
are generated.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from pathlib import Path
from chunkarena.config import BENCHMARK_XLSX, GOLDEN_DATASET_PATH


DARK_BLUE  = "1F3864"
MED_BLUE   = "2F5496"
HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
ALT_ROW    = "EBF3FB"


def hfont(size=11, bold=True, color=HEADER_FG):
    return Font(name="Arial", size=size, bold=bold, color=color)


def bfont(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)


def hfill(color=HEADER_BG):
    return PatternFill("solid", fgColor=color)


def cfill(color):
    return PatternFill("solid", fgColor=color)


def tborder():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def ca():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def la():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def hrow(ws, row_num, cols):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font = hfont(); c.fill = hfill()
        c.alignment = ca(); c.border = tborder()


def col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_workbook(summary_df, raw_df, final_k, method_params=None, stats_df=None):
    print("\nBuilding Excel report...")

    wb = Workbook()
    wb.remove(wb.active)

    _build_experiment_matrix_sheet(wb, summary_df, final_k, method_params or {}, stats_df)
    _build_raw_results_sheet(wb, raw_df)
    _build_heatmap_sheet(wb, summary_df)

    wb.save(BENCHMARK_XLSX)
    print(f"\nExcel report saved -> {BENCHMARK_XLSX}")


def _build_experiment_matrix_sheet(wb, summary_df, final_k, method_params, stats_df=None):
    ws = wb.create_sheet("Experiment Matrix")

    # Dataset info banner row
    _ds_name = Path(GOLDEN_DATASET_PATH).stem
    n_header_cols = 26  # total header columns
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_header_cols)
    c = ws.cell(row=1, column=1, value=f"Golden Dataset: {_ds_name}")
    c.font = Font(name="Arial", size=12, bold=True, color=DARK_BLUE)
    c.alignment = la()
    ws.row_dimensions[1].height = 26

    headers = [
        "Experiment ID", "Retriever Method", "Chunking Method",
        "Chunk Size", "Overlap (%)", "Top-K",
        # --- Chunk stats (from stats_df) ---
        "Num Chunks", "Avg Words", "Std Words",
        "Min Words", "Max Words", "Median Words",
        "Corpus Boundary", "Corpus Redundancy",
        # --- Retrieval metrics ---
        "Recall@K", "Precision@K", "MRR", "nDCG@K", "Hit Rate",
        "Context Relevance", "Faithfulness", "Answer Correctness",
        "Token Cost", "Latency (ms)", "Redundancy Score", "Notes",
    ]
    hrow(ws, 2, headers)
    ws.row_dimensions[2].height = 32

    # Build a method -> chunk stats lookup
    stats_lookup = {}
    if stats_df is not None:
        for _, srow in stats_df.iterrows():
            stats_lookup[srow["method"]] = srow

    df = summary_df.sort_values(
        ["technique", "composite_score"], ascending=[True, False]
    ).reset_index(drop=True)

    lat_median = df["avg_latency_ms"].median() if len(df) else 0.0
    tc_median  = df["avg_token_cost"].median() if len(df) else 0.0

    def make_note(row):
        parts = [str(row["composite_score_verdict"])]
        if lat_median > 0 and row["avg_latency_ms"] < lat_median * 0.7:
            parts.append("fast")
        elif lat_median > 0 and row["avg_latency_ms"] > lat_median * 1.3:
            parts.append("slow")
        if tc_median > 0 and row["avg_token_cost"] > tc_median * 1.3:
            parts.append("high token cost")
        if row["redundancy"] >= 0.6:
            parts.append("redundant")
        return ", ".join(parts)

    # Column indices for left-aligned text columns
    text_cols = {2, 3, len(headers)}  # Retriever Method, Chunking Method, Notes

    for r_idx, row in enumerate(df.itertuples(index=False), 3):
        exp_id = f"EXP-{r_idx - 2:02d}"
        params = method_params.get(row.method, {"chunk_size": "-", "overlap_pct": "-"})
        st = stats_lookup.get(row.method, {})

        vals = [
            exp_id,
            row.technique,
            row.method,
            params.get("chunk_size", "-"),
            params.get("overlap_pct", "-"),
            final_k,
            # Chunk stats
            int(st.get("num_chunks", 0)) if st is not None and len(st) else "-",
            round(st.get("avg_words", 0), 2) if st is not None and len(st) else "-",
            round(st.get("std_words", 0), 2) if st is not None and len(st) else "-",
            int(st.get("min_words", 0)) if st is not None and len(st) else "-",
            int(st.get("max_words", 0)) if st is not None and len(st) else "-",
            round(st.get("median_words", 0), 2) if st is not None and len(st) else "-",
            round(st.get("boundary_ratio", 0), 4) if st is not None and len(st) else "-",
            round(st.get("collection_redundancy", 0), 4) if st is not None and len(st) else "-",
            # Retrieval metrics
            round(row.recall_at_k, 4),
            round(row.precision_at_k, 4),
            round(row.mrr, 4),
            round(row.ndcg_at_k, 4),
            round(row.hit_at_k, 4),
            round(row.context_relevance, 4),
            round(row.faithfulness, 4),
            round(row.answer_correctness, 4),
            int(round(row.avg_token_cost)),
            round(row.avg_latency_ms, 2),
            round(row.redundancy, 4),
            make_note(row._asdict() if hasattr(row, "_asdict") else row.__dict__),
        ]

        alt = cfill(ALT_ROW) if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(vals, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = bfont(size=9, bold=(c_idx <= 3))
            c.alignment = la() if c_idx in text_cols else ca()
            c.border = tborder()
            if alt:
                c.fill = alt

    ws.freeze_panes = "D3"

    n_rows = len(df) + 2
    # Higher better: Recall(O), Precision(P), MRR(Q), nDCG(R), Hit(S),
    # Context Relevance(T), Faithfulness(U), Answer Correctness(V)
    for col_letter in ["O", "P", "Q", "R", "S", "T", "U", "V"]:
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{n_rows}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B")
        )
    # Lower better: Token Cost(W), Latency(X), Redundancy Score(Y)
    for col_letter in ["W", "X", "Y"]:
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{n_rows}",
            ColorScaleRule(start_type="min", start_color="63BE7B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="F8696B")
        )

    col_widths(ws, [
        12, 16, 16, 12, 12, 7,           # ID, Retriever, Method, Size, Overlap, K
        12, 11, 11, 11, 11, 13, 14, 16,  # Chunk stats
        10, 12, 9, 10, 10, 16, 13, 18,   # Retrieval metrics (Recall, Prec, MRR, nDCG, Hit, CR, Faith, AC)
        12, 14, 14, 32,                   # Cost, Latency, Redundancy, Notes
    ])


def _build_raw_results_sheet(wb, raw_df):
    ws = wb.create_sheet("Raw Results")
    ws.freeze_panes = "E2"

    raw_cols = ["Method", "Technique", "Q_ID", "Question",
                "Hit@K", "MRR", "Precision@K", "nDCG@K", "Recall@K",
                "Avg Rank", "Redundancy", "Boundary",
                "Context Relevance", "Faithfulness", "Answer Correctness",
                "Token Cost", "Latency (ms)"]
    hrow(ws, 1, raw_cols)

    raw_disp = raw_df[[
        "method", "technique", "question_id", "question",
        "hit@k", "mrr", "precision@k", "ndcg@k", "recall@k",
        "avg_rank", "redundancy", "boundary",
        "context_relevance", "faithfulness", "answer_correctness",
        "token_cost", "latency_ms"
    ]]

    for r_idx, row in enumerate(raw_disp.itertuples(index=False), 2):
        alt = cfill(ALT_ROW) if r_idx % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=ci, value=val)
            c.font = bfont(size=9)
            c.alignment = la() if ci == 4 else ca()
            c.border = tborder()
            if alt:
                c.fill = alt

    col_widths(ws, [14, 16, 6, 50, 8, 8, 12, 9, 10, 10, 12, 10, 16, 13, 18, 12, 14])
    ws.row_dimensions[1].height = 28


def _build_heatmap_sheet(wb, summary_df):
    ws = wb.create_sheet("Heatmap")

    metrics_heat = ["hit_at_k", "mrr", "precision_at_k", "ndcg_at_k",
                    "recall_at_k", "context_relevance", "faithfulness",
                    "answer_correctness", "redundancy", "composite_score"]
    m_labels_heat = ["Hit@K", "MRR", "Precision@K", "nDCG@K", "Recall@K",
                     "Context Relev.", "Faithfulness", "Answer Correct.",
                     "Redundancy", "Composite"]

    # All techniques × all methods, sorted by composite score (high → low)
    heat_data = summary_df.sort_values(
        "composite_score", ascending=False
    ).reset_index(drop=True)

    n_cols = 2 + len(m_labels_heat)  # Method + Technique + metrics
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws["A1"]
    _ds_name = Path(GOLDEN_DATASET_PATH).stem
    c.value = f"Performance Heatmap — {_ds_name} — All Techniques (sorted by Composite Score)"
    c.font = Font(name="Arial", size=13, bold=True, color=DARK_BLUE)
    c.alignment = ca()
    ws.row_dimensions[1].height = 28

    hrow(ws, 2, ["Method", "Technique"] + m_labels_heat)

    for r_idx, row in enumerate(heat_data.itertuples(index=False), 3):
        c = ws.cell(row=r_idx, column=1, value=row.method)
        c.font = bfont(bold=True); c.border = tborder(); c.alignment = la()
        c = ws.cell(row=r_idx, column=2, value=row.technique)
        c.font = bfont(); c.border = tborder(); c.alignment = la()
        for ci, met in enumerate(metrics_heat, 3):
            val = getattr(row, met)
            cell = ws.cell(row=r_idx, column=ci, value=round(val, 4))
            cell.font = bfont(); cell.alignment = ca(); cell.border = tborder()
        ws.row_dimensions[r_idx].height = 20

    n_heat = len(heat_data)
    data_start = 3
    data_end = n_heat + 2
    # Higher better: columns 3..10 (Hit through Answer Correct.) and column 12 (Composite)
    for ci in list(range(3, 11)) + [12]:
        cl = get_column_letter(ci)
        ws.conditional_formatting.add(
            f"{cl}{data_start}:{cl}{data_end}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B")
        )
    # Lower better: column 11 (Redundancy)
    cl = get_column_letter(11)
    ws.conditional_formatting.add(
        f"{cl}{data_start}:{cl}{data_end}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B")
    )

    col_widths(ws, [18, 16, 10, 10, 12, 10, 10, 14, 12, 16, 12, 12])
